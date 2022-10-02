from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from email.mime.multipart import MIMEMultipart
from selenium.webdriver.common.by import By
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from selenium import webdriver
from email import encoders
import openpyxl
import smtplib
import time
import json


#Configurando e Atualizando o webdriver
chrome = service=ChromeService(ChromeDriverManager().install())
options_chrome = webdriver.ChromeOptions()
options_chrome.add_argument('--start-maximized')
options_chrome.add_argument('--log-level=3')
driver = webdriver.Chrome(service=chrome, options=options_chrome)

class Scraping:

    def __init__(self,driver):
        driver = driver

        #Pagina de pesquisa
        self.searche = By.NAME, 'conteudo'
        self.btnSearche = By.XPATH, '//*[@id="rsyswpsdk"]/div/header/div[1]/div[1]/div/div[1]/form/button'
        
        #Raspagem de dados
        self.produto = By.XPATH, "*//div[@class='product-info__Container-sc-1or28up-0 cdKgxb']//h3"
        self.preco = By.XPATH, "*//span[@class='src__Text-sc-154pg0p-0 price__PromotionalPrice-sc-h6xgft-1 ctBJlj price-info__ListPriceWithMargin-sc-1xm1xzb-2 liXDNM']"
        self.btnprox = By.XPATH, '//*[@id="rsyswpsdk"]/div/main/div/div[3]/div[3]/div/ul/li[10]/button'

        #Listas das informações
        self.lista_descricao = []
        self.lista_preco = []

    def start(self):
        self.search_page()
        self.creat_worksheet()
        self.send_email()

    def search_page(self):
        last_Page = 'disabled=""'
        driver.get('https://www.americanas.com.br')

        driver.find_element(*self.searche).send_keys('placa de video gtx 1050ti')
        driver.find_element(*self.btnSearche).click()

        time.sleep(10)
        self.getting_information()
        while True:
            try:
                if last_Page in driver.find_element(By.XPATH, '//*[@id="rsyswpsdk"]/div/main/div/div[3]/div[3]/div/ul/li[10]').get_attribute('innerHTML'):
                    print('last page')
                    break
                else:
                    driver.find_element(*self.btnprox).click()
                    print(f'\u001b[35m{"Proxima pagina"}\u001b[0m')
                    time.sleep(3)
                    self.getting_information()
            except:
                time.sleep(1)
    
    def getting_information(self):
        quant = len(driver.find_elements(*self.produto))
        for i in range(quant): 
            produto = driver.find_elements(*self.produto)[i].text
            preco = driver.find_elements(*self.preco)[i].text

            self.lista_descricao.append(produto)
            self.lista_preco.append(preco)
    
    def creat_worksheet(self):
        index = 2
        wb = openpyxl.Workbook()
        ws = wb['Sheet']
        ws.title = 'produtos'
        ws['A1'] = 'Descrição'
        ws['B1'] = 'Preço'

        for nome, preco in zip (self.lista_descricao, self.lista_preco):
            ws.cell(column=1, row=index, value=nome)
            ws.cell(column=2, row=index, value=preco)
            index += 1

        wb.save('planilha_de_preços.xlsx')

        self.file = 'planilha_de_preços.xlsx'

        print(f'\u001b[32m{"Planilha criada com sucesso"}\u001b[0m')

    def send_email(self):

        with open('server.json', 'r') as f:
            server = json.load(f)
        login = server['login']
        password = server['password']

        server = smtplib.SMTP(server['host'], server['port'])
        server.ehlo()
        server.starttls()
        server.login(login, password)

        email_msg = MIMEMultipart()
        email_msg['From'] = login
        email_msg['To'] = login
        email_msg['Subject'] = 'Planilha de preços'

        email_msg.attach(MIMEText('Olá. <br> Segue anexo arquivo <b>.xlsx</b> com os nomes e preços correspondentes \
            a pesquisa do site: https://www.americanas.com.br', 'html'))

        attachment = open(self.file, 'rb')
        att = MIMEBase('application', 'octect-stream')
        att.set_payload(attachment.read())
        encoders.encode_base64(att)

        att.add_header('Content-Disposition', f'attachment; filename={self.file}')
        attachment.close()
        email_msg.attach(att)
        server.sendmail(email_msg['From'], email_msg['To'], email_msg.as_string())
        server.quit()
        print('\nEmail successfully sent.')


executar = Scraping(driver)
executar.start()
